"""Genera exports/TCPO_Paraguay_Completo_YYYYMMDD.xlsx con vistas limpias."""
import sys
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

import logging
logging.getLogger("streamlit").setLevel(logging.ERROR)
from app.utils.queries import get_conn

conn = get_conn()

# ── Estilo ───────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
FILL_REL = {
    "alta":           PatternFill("solid", fgColor="C6EFCE"),
    "media":          PatternFill("solid", fgColor="FFEB9C"),
    "baja":           PatternFill("solid", fgColor="FFCC99"),
    "no_aplica":      PatternFill("solid", fgColor="FFC7CE"),
    "sin_clasificar": PatternFill("solid", fgColor="D9D9D9"),
}
FONT_REL = {
    "alta":           Font(color="006100", bold=True),
    "media":          Font(color="9C6500"),
    "baja":           Font(color="7B3F00"),
    "no_aplica":      Font(color="9C0006"),
    "sin_clasificar": Font(color="595959"),
}

def aplicar_estilo(ws, col_widths: dict, rel_col: int = None):
    """Cabecera azul, freeze, autofilter y anchos de columna."""
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    if rel_col:
        for i, row in enumerate(ws.iter_rows(min_row=2, min_col=rel_col,
                                             max_col=rel_col), start=2):
            cell = row[0]
            val  = cell.value or "sin_clasificar"
            cell.fill = FILL_REL.get(val, FILL_REL["sin_clasificar"])
            cell.font = FONT_REL.get(val, FONT_REL["sin_clasificar"])
            cell.alignment = Alignment(horizontal="center")


# ── 1. Solo Servicios (SER.CG) — lo que va en un presupuesto ────────────────
print("Cargando Servicios SER.CG ...")
df_srv = pd.read_sql_query("""
    SELECT
        capitulo                                             AS Capitulo,
        COALESCE(subcapitulo, '(sin subcapitulo)')           AS Subcapitulo,
        codigo                                               AS Codigo,
        COALESCE(descripcion_es, descripcion_pt)             AS Descripcion,
        descripcion_pt                                       AS Descripcion_PT,
        CASE WHEN descripcion_es IS NOT NULL THEN 'Si' ELSE 'No' END AS Traducido,
        unidad                                               AS Unidad,
        precio_brl                                           AS Precio_BRL,
        precio_total_brl                                     AS Total_BRL,
        COALESCE(relevancia_py, 'sin_clasificar')            AS Relevancia_PY,
        COALESCE(relevancia_justificacion, '')               AS Justificacion
    FROM tcpo_items
    WHERE class = 'SER.CG'
    ORDER BY capitulo, COALESCE(subcapitulo,'zz'), codigo
""", conn)
print(f"  {len(df_srv):,} servicios unicos")


# ── 2. Insumos unicos (MAT / M.O. / EQ — sin repetir codigo) ────────────────
print("Cargando insumos unicos ...")
df_ins = pd.read_sql_query("""
    SELECT
        class                                                AS Clase,
        codigo                                               AS Codigo,
        COALESCE(descripcion_es, descripcion_pt)             AS Descripcion,
        descripcion_pt                                       AS Descripcion_PT,
        CASE WHEN descripcion_es IS NOT NULL THEN 'Si' ELSE 'No' END AS Traducido,
        unidad                                               AS Unidad,
        ROUND(coef, 4)                                       AS Coeficiente,
        precio_brl                                           AS Precio_Unitario_BRL,
        COUNT(*) OVER (PARTITION BY codigo)                  AS Veces_Usado
    FROM tcpo_items
    WHERE class IN ('MAT.','M.O.','EQ.AQ.','SER.CH','DIVER')
      AND id = (
            SELECT MIN(id) FROM tcpo_items t2
            WHERE t2.codigo = tcpo_items.codigo
      )
    ORDER BY class, codigo
""", conn)
print(f"  {len(df_ins):,} insumos unicos")


# ── 3. Composicion normalizada (relacion servicio → insumo + coeficiente) ────
# Cada fila es un par unico (servicio, insumo) con su coeficiente.
# Las descripciones de servicios estan en hoja 1; las de insumos en hoja 2.
# Esto evita repetir miles de veces el mismo insumo con distinto contexto.
print("Cargando composicion normalizada ...")
df_comp = pd.read_sql_query("""
    WITH servicios AS (
        -- Para cada insumo buscamos el SER.CG raiz que lo precede
        SELECT
            ins.id                                          AS ins_id,
            ins.codigo                                      AS Insumo_Codigo,
            ins.class                                       AS Clase,
            ROUND(ins.coef, 6)                              AS Coeficiente,
            ins.precio_brl                                  AS Precio_Unit_BRL,
            ROUND(ins.precio_total_brl, 4)                  AS Total_BRL,
            (
                SELECT srv.codigo
                FROM tcpo_items srv
                WHERE srv.class = 'SER.CG'
                  AND srv.coef  = 1.0
                  AND srv.id    = (
                        SELECT MAX(s2.id)
                        FROM tcpo_items s2
                        WHERE s2.class = 'SER.CG'
                          AND s2.coef  = 1.0
                          AND s2.id   <= ins.id
                  )
            )                                               AS Servicio_Codigo
        FROM tcpo_items ins
        WHERE ins.class != 'SER.CG'
           OR (ins.class = 'SER.CG' AND ins.coef != 1.0)
    )
    SELECT
        Servicio_Codigo,
        Insumo_Codigo,
        Clase,
        Coeficiente,
        Precio_Unit_BRL,
        Total_BRL
    FROM servicios
    WHERE Servicio_Codigo IS NOT NULL
    ORDER BY Servicio_Codigo, ins_id
""", conn)
print(f"  {len(df_comp):,} relaciones servicio-insumo")


# ── 4. Cobertura por capitulo ────────────────────────────────────────────────
df_cob = pd.read_sql_query("""
    SELECT
        capitulo                                                          AS Capitulo,
        COUNT(*)                                                          AS Total,
        SUM(CASE WHEN class='SER.CG' THEN 1 ELSE 0 END)                  AS Servicios,
        SUM(CASE WHEN class IN ('MAT.','M.O.','EQ.AQ.') THEN 1 ELSE 0 END) AS Insumos,
        SUM(CASE WHEN descripcion_es IS NOT NULL THEN 1 ELSE 0 END)      AS Traducidos,
        ROUND(SUM(CASE WHEN descripcion_es IS NOT NULL THEN 1.0 ELSE 0 END)
              * 100 / COUNT(*), 1)                                        AS Pct_Trad,
        SUM(CASE WHEN relevancia_py='alta'       THEN 1 ELSE 0 END)      AS Alta,
        SUM(CASE WHEN relevancia_py='media'      THEN 1 ELSE 0 END)      AS Media,
        SUM(CASE WHEN relevancia_py='baja'       THEN 1 ELSE 0 END)      AS Baja,
        SUM(CASE WHEN relevancia_py='no_aplica'  THEN 1 ELSE 0 END)      AS No_Aplica
    FROM tcpo_items
    WHERE class IS NOT NULL
    GROUP BY capitulo ORDER BY capitulo
""", conn)


# ── 5. Mandua ────────────────────────────────────────────────────────────────
df_mat = pd.read_sql_query("""
    SELECT seccion AS Seccion, descripcion AS Descripcion,
           unidad AS Unidad, precio_gs AS Precio_Gs, fuente_edicion AS Edicion
    FROM mandua_materiales ORDER BY seccion, descripcion
""", conn)
df_mo = pd.read_sql_query("""
    SELECT seccion AS Seccion, descripcion AS Descripcion,
           unidad AS Unidad, precio_gs AS Precio_Gs, fuente_edicion AS Edicion
    FROM mandua_mano_obra ORDER BY seccion, descripcion
""", conn)


# ── Escribir Excel ────────────────────────────────────────────────────────────
OUT = ROOT / "exports" / f"TCPO_Paraguay_Completo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
OUT.parent.mkdir(exist_ok=True)
print(f"\nEscribiendo {OUT} ...")

with pd.ExcelWriter(str(OUT), engine="openpyxl") as writer:

    # Hoja 1: Servicios (la principal para presupuestos)
    df_srv.to_excel(writer, sheet_name="1_Servicios_SER.CG", index=False)
    aplicar_estilo(writer.sheets["1_Servicios_SER.CG"],
                   {"A":55,"B":40,"C":22,"D":70,"E":70,"F":10,"G":9,
                    "H":13,"I":13,"J":16,"K":60},
                   rel_col=10)
    print("  1_Servicios_SER.CG OK")

    # Hoja 2: Insumos unicos
    df_ins.to_excel(writer, sheet_name="2_Insumos_Unicos", index=False)
    aplicar_estilo(writer.sheets["2_Insumos_Unicos"],
                   {"A":10,"B":22,"C":70,"D":70,"E":10,"F":9,"G":12,
                    "H":18,"I":14})
    print("  2_Insumos_Unicos OK")

    # Hoja 3: Composicion normalizada
    # Estructura relacional: Servicio_Codigo | Insumo_Codigo | Clase | Coef | Precio_Unit | Total
    # Las descripciones se buscan con VLOOKUP/BUSCARV en hojas 1 y 2.
    df_comp.to_excel(writer, sheet_name="3_Composicion", index=False)
    ws3 = writer.sheets["3_Composicion"]
    aplicar_estilo(ws3, {"A":22,"B":22,"C":10,"D":12,"E":16,"F":16})
    # Zebra por servicio: alternar fondo cada vez que cambia Servicio_Codigo
    FILL_A = PatternFill("solid", fgColor="F2F2F2")
    FILL_B = PatternFill("solid", fgColor="FFFFFF")
    last_srv, flip = None, False
    for i, srv in enumerate(df_comp["Servicio_Codigo"], start=2):
        if srv != last_srv:
            flip = not flip
            last_srv = srv
        fill = FILL_A if flip else FILL_B
        for col in range(1, 7):
            ws3.cell(row=i, column=col).fill = fill
    print("  3_Composicion OK")

    # Hoja 4: Cobertura
    df_cob.to_excel(writer, sheet_name="4_Cobertura", index=False)
    ws4 = writer.sheets["4_Cobertura"]
    aplicar_estilo(ws4, {"A":55,"B":14,"C":14,"D":14,"E":14,"F":14,
                         "G":14,"H":14,"I":14,"J":14,"K":14})
    for i in range(2, len(df_cob) + 2):
        pct = ws4.cell(row=i, column=6).value or 0
        ws4.cell(row=i, column=6).fill = (
            PatternFill("solid", fgColor="C6EFCE") if pct >= 95 else
            PatternFill("solid", fgColor="FFEB9C") if pct >= 50 else
            PatternFill("solid", fgColor="FFC7CE"))
    print("  4_Cobertura OK")

    # Hoja 5: Mandua materiales
    df_mat.to_excel(writer, sheet_name="5_Mandua_Materiales", index=False)
    aplicar_estilo(writer.sheets["5_Mandua_Materiales"],
                   {"A":30,"B":70,"C":10,"D":15,"E":20})
    print("  5_Mandua_Materiales OK")

    # Hoja 6: Mandua mano de obra
    df_mo.to_excel(writer, sheet_name="6_Mandua_Mano_Obra", index=False)
    aplicar_estilo(writer.sheets["6_Mandua_Mano_Obra"],
                   {"A":30,"B":70,"C":10,"D":15,"E":20})
    print("  6_Mandua_Mano_Obra OK")

size_mb = OUT.stat().st_size / 1024 / 1024
print(f"\nListo: {OUT}")
print(f"Tamano: {size_mb:.1f} MB")
