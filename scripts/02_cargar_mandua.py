"""02_cargar_mandua.py - Carga datos de Mandu'a Excel a SQLite."""
import sys
import re
from pathlib import Path
from datetime import date

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from unidecode import unidecode as _unidecode

from src.db import get_connection
from src.config import MANDUA_XLSX

FUENTE = "N°514 Mar 2026"
FECHA_CARGA = date.today().isoformat()

# Fill colors: ARGB hex strings as returned by openpyxl
_SECTION_COLORS = {"FFC8E6C9", "FFF1C40F"}       # green (Mano de Obra) / yellow (Materiales, Costeo)
_SUBSECTION_COLORS = {"FFEDE7F6", "FFE8DAEF"}     # light purple variants
_SUBSUBSECTION_COLORS = {"FFF5EEF8", "FFF3EEF6"}  # very light purple (item groups in Materiales)

EXPECTED = {
    "mandua_mano_obra":  (411, 511),   # 461 filas reales en N°514 Mar 2026
    "mandua_materiales": (451, 551),   # 501 filas reales
    "mandua_costeo":     (120, 220),
}


def normalizar_texto(s):
    if not s:
        return None
    s = _unidecode(str(s)).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def parse_precio(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.lower() in ("s/c", "", "-"):
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _cell_fill(cell):
    try:
        return cell.fill.fgColor.rgb
    except Exception:
        return "FFFFFFFF"


def _build_merged_index(ws):
    """Return {row_num: (value, fill_rgb)} for every single-row merge starting at col A."""
    idx = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.min_row == mr.max_row:
            cell = ws.cell(row=mr.min_row, column=1)
            idx[mr.min_row] = (cell.value, _cell_fill(cell))
    return idx


def _classify(fill):
    if fill in _SECTION_COLORS:
        return "section"
    if fill in _SUBSECTION_COLORS:
        return "subsection"
    if fill in _SUBSUBSECTION_COLORS:
        return "subsubsection"
    return "subsection"  # safe default for unknown merged fills


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _cargar_simple(ws, tabla, conn):
    """Load Mano de Obra or Materiales (col layout: A=desc, B=un, C=precio)."""
    merged = _build_merged_index(ws)
    seccion = subseccion = _base_subseccion = None
    rows = []
    warnings = []

    for row in ws.iter_rows(min_row=2):
        rn = row[0].row
        if rn in merged:
            val, fill = merged[rn]
            level = _classify(fill)
            text = str(val).strip() if val else None
            if level == "section":
                seccion = text
                subseccion = _base_subseccion = None
            elif level == "subsection":
                subseccion = _base_subseccion = text
            elif level == "subsubsection":
                # Each sub-subsection resets relative to the current base subseccion
                subseccion = f"{_base_subseccion} \u2013 {text}" if _base_subseccion else text
            continue

        desc = row[0].value
        if not desc or str(desc).strip() == "":
            continue
        desc = str(desc).strip()

        unidad = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
        precio_raw = row[2].value if len(row) > 2 else None

        if precio_raw is None:
            warnings.append(f"  WARNING precio vacío: '{desc}'")
        elif isinstance(precio_raw, str):
            s = precio_raw.strip().lower()
            if s not in ("s/c", "", "-"):
                warnings.append(f"  WARNING precio string inesperado '{precio_raw}': '{desc}'")

        rows.append((
            desc,
            normalizar_texto(desc),
            unidad,
            parse_precio(precio_raw),
            seccion,
            subseccion,
            FUENTE,
            FECHA_CARGA,
        ))

    conn.execute(f"DELETE FROM {tabla}")
    conn.executemany(
        f"INSERT INTO {tabla} "
        "(descripcion, descripcion_normalizada, unidad, precio_gs, "
        "seccion, subseccion, fuente_edicion, fecha_carga) "
        "VALUES (?,?,?,?,?,?,?,?)",
        rows,
    )
    for w in warnings:
        print(w)
    return len(rows)


def _cargar_costeo(ws, conn):
    """Load Costeo de Obra: save one row per 'Total por:' line."""
    merged = _build_merged_index(ws)
    seccion = partida = None
    rows = []

    for row in ws.iter_rows(min_row=2):
        rn = row[0].row
        if rn in merged:
            val, fill = merged[rn]
            level = _classify(fill)
            text = str(val).strip() if val else None
            if level == "section":
                seccion = text
                partida = None
            else:
                partida = text
            continue

        desc_a = row[0].value
        if not desc_a:
            continue
        if str(desc_a).strip().lower().startswith("total por"):
            unidad = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
            precio_raw = row[4].value if len(row) > 4 else None  # col E = SUBTOTAL
            if partida and precio_raw is not None:
                rows.append((
                    partida,
                    seccion,
                    unidad,
                    parse_precio(precio_raw),
                    FUENTE,
                    FECHA_CARGA,
                ))

    conn.execute("DELETE FROM mandua_costeo")
    conn.executemany(
        "INSERT INTO mandua_costeo "
        "(partida, seccion, unidad, precio_total_gs, fuente_edicion, fecha_carga) "
        "VALUES (?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Leyendo {MANDUA_XLSX} ...")
    wb = openpyxl.load_workbook(MANDUA_XLSX)
    conn = get_connection()

    with conn:
        n_mo  = _cargar_simple(wb["Mano de Obra"], "mandua_mano_obra", conn)
        n_mat = _cargar_simple(wb["Materiales"],   "mandua_materiales", conn)
        n_co  = _cargar_costeo(wb["Costeo de Obra"], conn)

    counts = {
        "mandua_mano_obra":  n_mo,
        "mandua_materiales": n_mat,
        "mandua_costeo":     n_co,
    }

    print("\n--- Resumen de carga ---")
    all_ok = True
    for tabla, n in counts.items():
        lo, hi = EXPECTED[tabla]
        if lo <= n <= hi:
            status = "OK"
        else:
            status = f"WARNING: esperado {lo}-{hi}"
            all_ok = False
        print(f"  {tabla:<25} {n:>4} filas  [{status}]")
    if all_ok:
        print("\nTodas las tablas dentro del rango esperado.")

    print("\n--- Muestras (3 filas c/u) ---")
    for tabla in ("mandua_mano_obra", "mandua_materiales", "mandua_costeo"):
        print(f"\n{tabla}:")
        for row in conn.execute(f"SELECT * FROM {tabla} LIMIT 3"):
            print(dict(row))

    conn.close()


if __name__ == "__main__":
    main()
